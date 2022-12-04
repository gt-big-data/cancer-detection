import openpyxl

wrkbk = openpyxl.load_workbook("cancer1.xlsx")

sh = wrkbk.active

for i in range(1, sh.max_row+1):
    print("\n")
    print("Row ", i, " data :")

    for j in range(1):
        cell_obj = sh.cell(row=i, column=1)
        cell_obj1 = sh.cell(row = i, column = 10)
        cell_obj2 = sh.cell(row = i, column = 14)
    
        print(cell_obj.value + " " + cell_obj1.value + " " + cell_obj2.value) 